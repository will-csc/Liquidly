from flask import Flask, request, jsonify
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
import random
import threading
import time
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from io import BytesIO
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()

app = Flask(__name__)

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USERNAME = (os.getenv("SMTP_USERNAME") or "").strip() or None
_raw_smtp_password = os.getenv("SMTP_PASSWORD") or ""
SMTP_PASSWORD = "".join(_raw_smtp_password.split()) or None
API_KEY = os.getenv("EMAIL_SERVICE_API_KEY")
SMTP_USE_SSL = os.getenv("SMTP_USE_SSL", "").lower() in ("1", "true", "yes", "on") or SMTP_PORT == 465
SMTP_USE_STARTTLS = os.getenv("SMTP_USE_STARTTLS", "true").lower() in ("1", "true", "yes", "on")
print(f"SMTP config: server={SMTP_SERVER} port={SMTP_PORT} user={SMTP_USERNAME} ssl={SMTP_USE_SSL} starttls={SMTP_USE_STARTTLS} password_set={bool(SMTP_PASSWORD)}")

GENERATOR_ENABLED = os.getenv("GENERATOR_ENABLED", "false").lower() in ("1", "true", "yes", "on")
GENERATOR_INTERVAL_SECONDS = int(os.getenv("GENERATOR_INTERVAL_SECONDS", "60"))
GENERATOR_MIN_ROWS = int(os.getenv("GENERATOR_MIN_ROWS", "1"))
GENERATOR_MAX_ROWS = int(os.getenv("GENERATOR_MAX_ROWS", "10"))
DATABASE_URL = os.getenv("DATABASE_URL")
DB_SSLMODE = os.getenv("DB_SSLMODE", "prefer")

_generator_thread = None
_generator_stop = threading.Event()

def _check_api_key():
    if not API_KEY:
        return None
    provided = request.headers.get("X-API-Key")
    if provided != API_KEY:
        return jsonify({"error": "Unauthorized"}), 401
    return None

def _validate_smtp_config():
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        return jsonify({"error": "SMTP is not configured"}), 500
    return None

def _send_email(to_email: str, subject: str, body: str):
    msg = MIMEMultipart()
    msg["From"] = SMTP_USERNAME
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    if SMTP_USE_SSL:
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)
        return

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        if SMTP_USE_STARTTLS:
            server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(msg)

def _send_email_with_attachment(to_email: str, subject: str, body: str, attachment_bytes: bytes, filename: str):
    msg = MIMEMultipart()
    msg["From"] = SMTP_USERNAME
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    if SMTP_USE_SSL:
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)
        return

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        if SMTP_USE_STARTTLS:
            server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(msg)

def _parse_application_properties(path: Path):
    if not path.exists():
        return {}
    data = {}
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        data[k.strip()] = v.strip()
    return data

def _build_database_url_fallback():
    props_path = Path(__file__).resolve().parents[2] / "backend-api" / "src" / "main" / "resources" / "application.properties"
    props = _parse_application_properties(props_path)
    jdbc_url = props.get("app.datasource.backup.url")
    username = props.get("app.datasource.backup.username")
    password = props.get("app.datasource.backup.password")

    if not jdbc_url or not username or not password:
        return None

    if not jdbc_url.startswith("jdbc:postgresql://"):
        return None

    without_prefix = jdbc_url.replace("jdbc:postgresql://", "", 1)
    if "/" not in without_prefix:
        return None

    host_port, dbname = without_prefix.split("/", 1)
    if ":" in host_port:
        host, port = host_port.split(":", 1)
    else:
        host, port = host_port, "5432"

    return f"postgresql://{username}:{password}@{host}:{port}/{dbname}"

def _get_db_conn():
    url = DATABASE_URL or _build_database_url_fallback()
    if not url:
        raise RuntimeError("DATABASE_URL is not configured")
    return psycopg2.connect(url, connect_timeout=5, sslmode=DB_SSLMODE)

def _quantize_4(value: Decimal):
    return value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

def _random_qty(max_qty: Decimal):
    if max_qty <= Decimal("0"):
        return Decimal("0")
    upper = max(Decimal("0.0001"), max_qty)
    raw = Decimal(str(random.uniform(0.01, float(upper))))
    return _quantize_4(raw)

def _eligible_company_project_pairs(cur):
    cur.execute(
        """
        SELECT DISTINCT b.company_id, b.project_id
        FROM boms b
        JOIN projects p ON p.id = b.project_id AND p.company_id = b.company_id
        JOIN users u ON u.company_id = b.company_id
        WHERE b.company_id IS NOT NULL AND b.project_id IS NOT NULL
        """
    )
    return cur.fetchall()

def _random_bom_for_project(cur, company_id: int, project_id: int):
    cur.execute(
        """
        SELECT item_code, item_name, qntd, um_bom, remaining_qntd
        FROM boms
        WHERE company_id = %s AND project_id = %s
        """,
        (company_id, project_id),
    )
    rows = cur.fetchall()
    if not rows:
        return None
    return random.choice(rows)

def _insert_invoice_and_po(cur, company_id: int, project_id: int):
    bom = _random_bom_for_project(cur, company_id, project_id)
    if not bom:
        return None

    item_code = str(bom["item_code"]).strip()
    um = bom["um_bom"]
    base_qty = bom["remaining_qntd"] if bom["remaining_qntd"] is not None else bom["qntd"]
    base_qty = Decimal(str(base_qty)) if base_qty is not None else Decimal("10")
    qty = _random_qty(max_qty=max(base_qty, Decimal("0.01")))
    unit_price = Decimal(str(random.uniform(10, 200))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    invoice_value = (qty * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    country = "BR"
    invoice_date_string = datetime.now().strftime("%Y-%m-%d")

    invoice_number = f"INV-{int(time.time())}-{random.randint(1000, 9999)}"

    cur.execute(
        """
        INSERT INTO invoices (project_id, item_code, invoice_number, country, invoice_date_string, invoice_value, qntd_invoice, um_invoice, remaining_qntd, company_id, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        RETURNING id
        """,
        (project_id, item_code, invoice_number, country, invoice_date_string, invoice_value, qty, um, qty, company_id),
    )
    invoice_id = cur.fetchone()["id"]

    cur.execute(
        """
        INSERT INTO pos (invoice_number, qntd_invoice, um_po, remaining_qntd, company_id, created_at)
        VALUES (%s, %s, %s, %s, %s, NOW())
        RETURNING id
        """,
        (invoice_number, qty, um, qty, company_id),
    )
    po_id = cur.fetchone()["id"]

    return {"invoiceId": invoice_id, "poId": po_id, "invoiceNumber": invoice_number}

def _parse_iso_date(value):
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date()
    except Exception:
        return None

def _build_liquidation_report_excel(rows, filters):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = [
        "project_name",
        "item",
        "um_bom",
        "qntd_bom",
        "qntd_consumed_bom",
        "remaining_qntd",
        "invoice_number",
        "invoice_country",
        "invoice_date_string",
        "invoice_value",
        "um_invoice",
        "qntd_invoice",
        "consumed_invoice_value",
        "qntd_consumed_invoice",
        "remaining_invoice_value",
        "remaining_qntd_invoice",
        "po_number",
        "um_po",
        "qntd_po",
        "qntd_consumed_po",
        "remaining_qntd_po",
        "created_at",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    numeric_cols = {
        "qntd_bom",
        "qntd_consumed_bom",
        "remaining_qntd",
        "qntd_invoice",
        "qntd_consumed_invoice",
        "remaining_qntd_invoice",
        "qntd_po",
        "qntd_consumed_po",
        "remaining_qntd_po",
    }
    money_cols = {
        "invoice_value",
        "consumed_invoice_value",
        "remaining_invoice_value",
    }
    for idx, h in enumerate(headers, start=1):
        if h in numeric_cols or h in money_cols:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=idx)
                if cell.value is None:
                    continue
                cell.number_format = "0.00" if h in money_cols else "0.0000"

    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                l = 0
            else:
                l = len(str(v))
            widths[i] = max(widths.get(i, 0), min(55, l))
    for i, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(10, w + 2)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def _fetch_liquidation_rows(company_id: int, project_id: int, item_code: str = None, start_date=None, end_date=None):
    conn = _get_db_conn()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                params = [company_id, project_id]
                where = ["lr.company_id = %s", "lr.project_id = %s"]

                if item_code:
                    where.append("lower(trim(lr.item)) = lower(trim(%s))")
                    params.append(item_code)

                if start_date and end_date:
                    where.append(
                        "COALESCE(i.created_at, p.created_at, lr.created_at) >= %s AND COALESCE(i.created_at, p.created_at, lr.created_at) < (%s::date + INTERVAL '1 day')"
                    )
                    params.append(start_date)
                    params.append(end_date)

                query = f"""
                    SELECT
                        lr.project_name,
                        lr.item,
                        lr.um_bom,
                        lr.qntd_bom,
                        lr.qntd_consumed_bom,
                        lr.remaining_qntd,
                        lr.invoice_number,
                        lr.invoice_country,
                        lr.invoice_date_string,
                        lr.invoice_value,
                        lr.um_invoice,
                        lr.qntd_invoice,
                        lr.consumed_invoice_value,
                        lr.qntd_consumed_invoice,
                        lr.remaining_invoice_value,
                        lr.remaining_qntd_invoice,
                        lr.po_number,
                        lr.um_po,
                        lr.qntd_po,
                        lr.qntd_consumed_po,
                        lr.remaining_qntd_po,
                        lr.created_at
                    FROM liquidation_results lr
                    LEFT JOIN invoices i
                        ON i.company_id = lr.company_id
                       AND i.project_id = lr.project_id
                       AND i.invoice_number = lr.invoice_number
                    LEFT JOIN pos p
                        ON p.company_id = lr.company_id
                       AND p.invoice_number = lr.po_number
                    WHERE {" AND ".join(where)}
                    ORDER BY lr.created_at NULLS LAST, lr.id
                """
                cur.execute(query, tuple(params))
                return cur.fetchall()
    finally:
        conn.close()

def _resolve_item_code(cur, company_id: int, bom_id=None, item_code=None):
    if item_code:
        return str(item_code).strip() or None
    if bom_id is None:
        return None
    try:
        bom_id_int = int(bom_id)
    except Exception:
        return None
    cur.execute(
        "SELECT item_code FROM boms WHERE company_id = %s AND id = %s LIMIT 1",
        (company_id, bom_id_int),
    )
    row = cur.fetchone()
    if not row:
        return None
    return row["item_code"]

def generate_invoices_and_pos(min_rows: int = None, max_rows: int = None):
    min_rows = GENERATOR_MIN_ROWS if min_rows is None else int(min_rows)
    max_rows = GENERATOR_MAX_ROWS if max_rows is None else int(max_rows)
    if min_rows < 0 or max_rows < 0 or min_rows > max_rows:
        raise RuntimeError("Invalid row range")

    target = random.randint(min_rows, max_rows)
    created = []

    conn = _get_db_conn()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                pairs = _eligible_company_project_pairs(cur)
                if not pairs:
                    return {"created": [], "requested": target}

                for _ in range(target):
                    picked = random.choice(pairs)
                    if isinstance(picked, dict):
                        company_id = picked["company_id"]
                        project_id = picked["project_id"]
                    else:
                        company_id, project_id = picked
                    res = _insert_invoice_and_po(cur, int(company_id), int(project_id))
                    if res:
                        created.append(res)
        return {"created": created, "requested": target}
    finally:
        conn.close()

def _generator_loop():
    while not _generator_stop.is_set():
        try:
            generate_invoices_and_pos()
        except Exception as e:
            print(f"Generator error: {type(e).__name__}: {e}")
        _generator_stop.wait(GENERATOR_INTERVAL_SECONDS)

def _ensure_generator_started():
    global _generator_thread
    if _generator_thread and _generator_thread.is_alive():
        return False
    _generator_stop.clear()
    _generator_thread = threading.Thread(target=_generator_loop, daemon=True)
    _generator_thread.start()
    return True

def _ensure_generator_stopped():
    if not _generator_thread:
        return False
    _generator_stop.set()
    return True

@app.route('/send-email', methods=['POST'])
def send_email():
    try:
        api_check = _check_api_key()
        if api_check:
            return api_check
        smtp_check = _validate_smtp_config()
        if smtp_check:
            return smtp_check

        data = request.get_json(silent=True)
        
        if not data or 'to' not in data or 'subject' not in data or 'body' not in data:
            return jsonify({"error": "Fields 'to', 'subject' and 'body' are required"}), 400

        to_email = data['to']
        subject = data['subject']
        body = data['body']

        _send_email(to_email=to_email, subject=subject, body=body)

        return jsonify({"message": "Email sent successfully"}), 200

    except Exception as e:
        print(f"Error sending email: {type(e).__name__}: {e}")
        return jsonify({"error": "Failed to send email"}), 500

@app.route('/send-report', methods=['POST'])
def send_report():
    try:
        api_check = _check_api_key()
        if api_check:
            return api_check
        smtp_check = _validate_smtp_config()
        if smtp_check:
            return smtp_check

        data = request.get_json(silent=True) or {}
        to_email = str(data.get("to") or "").strip()
        company_id = data.get("companyId")
        project_id = data.get("projectId")

        if not to_email or company_id is None or project_id is None:
            return jsonify({"error": "Fields 'to', 'companyId' and 'projectId' are required"}), 400

        try:
            company_id = int(company_id)
            project_id = int(project_id)
        except Exception:
            return jsonify({"error": "companyId and projectId must be integers"}), 400

        start_date = _parse_iso_date(data.get("startDate"))
        end_date = _parse_iso_date(data.get("endDate"))
        bom_id = data.get("bomId")
        item_code = data.get("itemCode")

        resolved_item_code = None
        conn = _get_db_conn()
        try:
            with conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                    resolved_item_code = _resolve_item_code(cur, company_id, bom_id=bom_id, item_code=item_code)
        finally:
            conn.close()

        rows = _fetch_liquidation_rows(
            company_id=company_id,
            project_id=project_id,
            item_code=resolved_item_code,
            start_date=start_date,
            end_date=end_date,
        )

        filters = {
            "companyId": company_id,
            "projectId": project_id,
            "itemCode": resolved_item_code,
            "startDate": str(start_date) if start_date else None,
            "endDate": str(end_date) if end_date else None,
        }

        attachment = _build_liquidation_report_excel(rows, filters)
        subject = str(data.get("subject") or "Liquidly Report").strip()

        body_lines = ["Seu relatório foi gerado com sucesso."]
        if filters["startDate"] and filters["endDate"]:
            body_lines.append(f"Período: {filters['startDate']} até {filters['endDate']}")
        if filters["itemCode"]:
            body_lines.append(f"Item: {filters['itemCode']}")
        body = "\n".join(body_lines)

        filename = str(data.get("filename") or "liquidly_report.xlsx").strip() or "liquidly_report.xlsx"
        _send_email_with_attachment(to_email=to_email, subject=subject, body=body, attachment_bytes=attachment, filename=filename)

        return jsonify({"message": "Report sent successfully", "rows": len(rows)}), 200
    except Exception as e:
        print(f"Error sending report: {type(e).__name__}: {e}")
        return jsonify({"error": "Failed to send report"}), 500

@app.route('/send-recovery-code', methods=['POST'])
def send_recovery_code():
    try:
        api_check = _check_api_key()
        if api_check:
            return api_check
        smtp_check = _validate_smtp_config()
        if smtp_check:
            return smtp_check

        data = request.get_json(silent=True)
        if not data or "email" not in data or "code" not in data:
            return jsonify({"error": "Fields 'email' and 'code' are required"}), 400

        to_email = str(data["email"]).strip()
        code = str(data["code"]).strip()

        if not to_email or not code:
            return jsonify({"error": "Fields 'email' and 'code' are required"}), 400

        subject = "Password recovery code"
        body = f"Your recovery code is: {code}"

        _send_email(to_email=to_email, subject=subject, body=body)
        return jsonify({"message": "Recovery code sent"}), 200
    except Exception as e:
        print(f"Error sending recovery code: {type(e).__name__}: {e}")
        return jsonify({"error": "Failed to send recovery code"}), 500

@app.route("/seed/invoices-pos", methods=["POST"])
def seed_invoices_pos():
    try:
        api_check = _check_api_key()
        if api_check:
            return api_check

        payload = request.get_json(silent=True) or {}
        result = generate_invoices_and_pos(
            min_rows=payload.get("minRows"),
            max_rows=payload.get("maxRows"),
        )
        return jsonify(result), 200
    except Exception as e:
        print(f"Seed error: {type(e).__name__}: {e}")
        return jsonify({"error": "Failed to seed invoices/pos"}), 500

@app.route("/seed/status", methods=["GET"])
def seed_status():
    return jsonify(
        {
            "generatorEnabled": GENERATOR_ENABLED,
            "generatorRunning": bool(_generator_thread and _generator_thread.is_alive()),
            "intervalSeconds": GENERATOR_INTERVAL_SECONDS,
        }
    ), 200

@app.route("/seed/start", methods=["POST"])
def seed_start():
    api_check = _check_api_key()
    if api_check:
        return api_check
    started = _ensure_generator_started()
    return jsonify({"started": started}), 200

@app.route("/seed/stop", methods=["POST"])
def seed_stop():
    api_check = _check_api_key()
    if api_check:
        return api_check
    stopped = _ensure_generator_stopped()
    return jsonify({"stopped": stopped}), 200

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy"}), 200

if GENERATOR_ENABLED:
    try:
        _ensure_generator_started()
    except Exception:
        print("Failed to start generator")

if __name__ == '__main__':
    port = int(os.getenv("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
